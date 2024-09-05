import abc

from google.cloud import speech, language_v1
from openpyxl import load_workbook
from pysondb import db
from .utils import is_empty
from werkzeug.utils import secure_filename

# Base Model
class Model(abc.ABC):
    _db_name = ""
    pk = "id"
    
    
    def __init__(self):
        self.db = self.get_db()
        
    @property
    def db_name(cls):
        return cls._db_name
    
    @classmethod
    def get_db(cls):
        return db.getDb(f"dbs/{cls._db_name}.json")
    
    def save(self, overwrite=False):
        obj = self.to_dict()
        exists = self.find(obj[self.pk])
        if exists:
            if not overwrite:# Object exists by unique value
                return False
            else:
                return self.db.update(exists, obj)
        
        return self.db.add(obj)

    def find(self, value):
        return self.db.getBy({
            self.pk: value 
        })
    
    def to_dict(self):
        raise NotImplementedError("Subclasses should implement the to_dict method")
        
class FileModel(Model, abc.ABC):
    pk = "filename"

    def __init__(self, file):
        super().__init__()
        self.file = file
        self.filename = secure_filename(file.filename)
        
    def find(self, filename):
        return self.db.getBy({
            self.pk: secure_filename(filename)
        })
    
class Audio(FileModel):
    _db_name = "audios"
    transcription = None
    response = None
    
    def __init__(self, file, do_transcribe=False):
        self.speech = speech.SpeechClient()
        super().__init__(file)
        
        if do_transcribe:
            self.transcribe(file)
    
    def transcribe(self, file): # We may want to control the user file that is being sent to the API.
        audio = self.find(file.filename)
        if audio:
            self.transcription = audio[0]["transcription"]
            return self.transcription

        if (not self.transcription):
            self.response = self.speech.recognize(
                config=speech.RecognitionConfig(
                    encoding=speech.RecognitionConfig.AudioEncoding.MP3,
                    sample_rate_hertz=16000,
                    language_code="en-US",
                    audio_channel_count=1,
                    model="default"
                ),
                audio=speech.RecognitionAudio(content=file.read())
            ).results
            
            if len(self.response) > 0:
                self.transcription = self.response[0].alternatives[0].transcript
                
        return self.transcription
        
    def to_dict(self):
        return {
            "filename": self.filename,
            "transcription": self.transcription
        }

class Excel(FileModel):
    _db_name = "excels"
    fields = [] # [str]
    wb = None
    sheet = None
    
    def __init__(self, file):
        super().__init__(file)
        self.wb = load_workbook(file)
        self.sheet = self.wb.active
        self.fields = self.get_fields(file)
        
    def get_fields(self, excel):
        first_row = self.get_first_row()

        if(self.sheet.max_row != 1) or is_empty(first_row):
            raise ValueError("Please make sure the first row of the excel is not empty and there are no other rows.")

        return first_row
    
    def get_first_row(self):
        return [cell.value for cell in self.sheet[1]]
    
    def format_content(self):
        headers = self.get_first_row()

        formatted = []
        for i, col in enumerate(self.sheet.iter_cols(min_row=2, values_only=True)):  # Start from the second row
            formatted.append({
                'header': headers[i],
                'values': [val for val in col if val is not None]
            })

        return formatted
    
    def populate_fields(self, sets):
        for col_i, (header, values) in enumerate(sets.items(), start=1):
            for row_i, value in enumerate(values, start=self.sheet.max_row + 1):
                self.sheet.cell(row=row_i, column=col_i, value=value)
        
    def to_dict(self):
        return {
            "filename": self.filename,
            "fields": self.fields
        }
        
class Analysis(Model):
    _db_name = "analysis"
    
    def __init__(self, audio_filename, excel_filename, field, value):
        self.audio_filename = audio_filename
        self.excel_filename = excel_filename
        self.field = field
        self.value = value
    
    @classmethod    
    def process(cls, audio_filename, excel_filename, results):
        return [cls(audio_filename, excel_filename, field, value) for field, value in results.items()]
    
    def to_dict(self):
        return {
            "audio_filename": self.audio_filename,
            "excel_filename": self.excel_filename,
            "field": self.field,
            "value": self.value
        }
        
    
